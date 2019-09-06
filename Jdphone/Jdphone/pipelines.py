# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html
import openpyxl


class JdphonePipeline(object):
    def open_spider(self, spider):  # 在爬虫开启的时候仅执行一次
        if spider.name == 'jdphone':
            self.name_list = []  # 手机去重列表
            # 没有过滤的2019手机信息表格
            self.f = openpyxl.Workbook()
            self.sheet1 = self.f.create_sheet("手机信息")
            self.sheet1.append({"A": "品牌", "B": "型号", "C": "价格", "D": "上市时间", "E": "好评率", "F": "图片", "G": "详情链接",
                                "H": "机身长度", "I": "机身宽度", "J": "机身重量", "K": "屏幕尺寸"})
            # 按照时间过滤后的信息表格
            self.filted_list = []  # 过滤出来的2019手机字典列表
            self.five_list = []  # 保存价格在500以下的手机
            self.onefive_list = []  # 保存价格在1500以下的手机
            self.twofive_list = []  # 保存价格在2500以下的手机
            self.thrfive_list = []  # 保存价格在3500以下的手机
            self.forfive_list = []  # 保存价格在3500以上的手机
            self.f1 = openpyxl.Workbook()
            self.sheet1 = self.f1.create_sheet("手机信息")
            self.sheet1.append({"A": "品牌", "B": "型号", "C": "价格", "D": "上市时间", "E": "好评率", "F": "图片", "G": "详情链接",
                                "H": "机身长度", "I": "机身宽度", "J": "机身重量", "K": "屏幕尺寸"})

    def close_spider(self, spider):  # 在爬虫关闭的时候仅执行一次
        if spider.name == 'jdphone':
            # 保存2019没有过滤的手机列表到表格中
            self.f.save("2019手机详情1.xlsx")
            # 遍历手机字典列表然后存储到不同价格区间的列表中


    def process_item(self, item, spider):
        print("管道开始执行了0000000000000000000000000000000000000000")
        # 过滤出2019年才发布的手机
        if item['year_time'] and item['year_time'][:4:] == "2019" and item['month_time'] != '以官网信息为准':
            # 再过滤出有名字的手机
            if item['name'] and item['name'] not in ['以官网信息为准', '·', '产品名称', '-', '--', '其他', '以官网数据为准', '官网信息为准',
                                                     '以官网为准']:
                # 再过滤出已经查询出来的手机
                if item['name'].upper().replace(" ", "") not in self.name_list:
                    # 再进行双向过滤出已经存在的手机
                    uppname = item['name'].upper().replace(" ", "")
                    if all(uppname not in element.upper().replace(" ", "") and element.upper().replace(" ",
                                                                                                       "") not in uppname
                           for element in self.name_list):
                        # 1.将过滤后的手机信息保存到表格中
                        self.sheet1.append(
                            {"A": item['brand'], "B": item['name'], "C": item['price'],
                             "D": item['year_time'] + item['month_time'],
                             "E": item['goodrate'], "F": item['image'], "G": item['link'], "H": item["length"],
                             "I": item["width"],
                             "J": item["weight"], "K": item["inch"]})
                        self.name_list.append(item['name'].upper().replace(" ", ""))
                        # 将过滤后的手机保存到字典列表中，
                        self.filted_list.append(item)
                        # 2.按照手机价格进行汇聚，然后按照时间进行排序

                        # 遍历出手机信，按照价格分别存储到不同的列表中


        print("管道执行结束=============================================")